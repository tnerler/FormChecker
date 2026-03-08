from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO


# ── Color palette ─────────────────────────────────────────────────────────────
CLR_HEADER_BG   = "1E3A5F"
CLR_HEADER_FG   = "FFFFFF"
CLR_SECTION_BG  = "2E4A6F"
CLR_SECTION_FG  = "FFFFFF"
CLR_TRUE_BG     = "D6F0E0"
CLR_TRUE_FG     = "1A6B3A"
CLR_FALSE_BG    = "FAD7D7"
CLR_FALSE_FG    = "8B0000"
CLR_NEUTRAL_BG  = "F5F5F5"
CLR_ROW_ALT     = "EEF3FA"
CLR_ROW_WHITE   = "FFFFFF"
CLR_BORDER      = "CCCCCC"

THIN_BORDER = Border(
    left=Side(style="thin", color=CLR_BORDER),
    right=Side(style="thin", color=CLR_BORDER),
    top=Side(style="thin", color=CLR_BORDER),
    bottom=Side(style="thin", color=CLR_BORDER),
)


def _cell(ws, row, col, value, bold=False, bg=None, fg="000000",
          wrap=False, center=False, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=fg, size=10)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    ha = "center" if center else "left"
    c.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    if border:
        c.border = THIN_BORDER
    return c


def _header_row(ws, row, columns: list):
    for col, label in enumerate(columns, 1):
        _cell(ws, row, col, label, bold=True,
              bg=CLR_HEADER_BG, fg=CLR_HEADER_FG, center=True)


def _section_row(ws, row, label: str, ncols: int):
    _cell(ws, row, 1, label, bold=True,
          bg=CLR_SECTION_BG, fg=CLR_SECTION_FG)
    for col in range(2, ncols + 1):
        _cell(ws, row, col, "", bg=CLR_SECTION_BG)
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)


def _bool_cell(ws, row, col, value):
    if isinstance(value, bool):
        text = "✔  YES" if value else "✘  NO"
        bg   = CLR_TRUE_BG if value else CLR_FALSE_BG
        fg   = CLR_TRUE_FG if value else CLR_FALSE_FG
        _cell(ws, row, col, text, bold=True, bg=bg, fg=fg, center=True)
    else:
        _cell(ws, row, col, str(value) if value else "—", center=True)


def _data_row(ws, row, label, value, is_bool=False, alt=False):
    bg = CLR_ROW_ALT if alt else CLR_ROW_WHITE
    _cell(ws, row, 1, label, bold=True, bg=bg)
    if is_bool:
        _bool_cell(ws, row, 2, value)
        ws.cell(row=row, column=2).fill = PatternFill(
            "solid",
            start_color=(CLR_TRUE_BG if value else CLR_FALSE_BG)
        )
    else:
        _cell(ws, row, 2, str(value) if value not in (None, "") else "—",
              bg=bg, wrap=True)


# ── Per-form-type sheet builders ─────────────────────────────────────────────

def _build_3a_sheet(ws, results: dict):
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 50

    _header_row(ws, 1, ["Field", "Value"])
    r = 2

    for form_name, data in results.items():
        _section_row(ws, r, f"📄  {form_name}", 2)
        r += 1

        fields = [
            ("Corresponding Author Exists",  data.get("corresponding_author_exists"), True),
            ("Corresponding Author Name",    data.get("corresponding_author_name"),   False),
            ("Title in Section 1 Exists",    data.get("title_in_1_exists"),           True),
            ("Title in Section 1",           data.get("title_in_1_name"),             False),
            ("Title in Section 2 Exists",    data.get("title_in_2_exists"),           True),
            ("Title in Section 2",           data.get("title_in_2_name"),             False),
            ("Signature Exists",             data.get("signature_exists"),            True),
            ("Date Exists",                  data.get("date_exists"),                 True),
            ("Date Value",                   data.get("date_value"),                  False),
            ("Name Exists",                  data.get("name_exists"),                 True),
            ("Name",                         data.get("name_value"),                  False),
            ("Surname Exists",               data.get("surname_exists"),              True),
            ("Surname",                      data.get("surname_value"),               False),
        ]
        for i, (label, value, is_bool) in enumerate(fields):
            _data_row(ws, r, label, value, is_bool=is_bool, alt=(i % 2 == 0))
            r += 1
        r += 1  # blank spacer


def _build_3b_sheet(ws, results: dict):
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 50

    _header_row(ws, 1, ["Field", "Value"])
    r = 2

    for form_name, data in results.items():
        _section_row(ws, r, f"📄  {form_name}", 2)
        r += 1

        fields = [
            ("Paper Title Exists",  data.get("paper_title_exists"), True),
            ("Paper Title",         data.get("paper_title_name"),   False),
            ("Signature Exists",    data.get("signature_exists"),   True),
            ("Date Exists",         data.get("date_exists"),        True),
            ("Date Value",          data.get("date_value"),         False),
            ("Name Exists",         data.get("name_exists"),        True),
            ("Name",                data.get("name_value"),         False),
            ("Surname Exists",      data.get("surname_exists"),     True),
            ("Surname",             data.get("surname_value"),      False),
        ]
        for i, (label, value, is_bool) in enumerate(fields):
            _data_row(ws, r, label, value, is_bool=is_bool, alt=(i % 2 == 0))
            r += 1
        r += 1


def _build_3c_sheet(ws, results: dict):
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 60

    _header_row(ws, 1, ["Field", "Value"])
    r = 2

    for form_name, data in results.items():
        _section_row(ws, r, f"📄  {form_name}", 2)
        r += 1

        authors = data.get("signature_of_authors", [])
        fields = [
            ("Paper Title Exists",      data.get("paper_title_exists"),        True),
            ("Paper Title",             data.get("paper_title_name"),           False),
            ("Signatures Exist",        data.get("signatures_exist"),           True),
            ("Signatures of Authors",   ", ".join(authors) if authors else "—", False),
            ("Co-Author Name Exists",   data.get("co_author_name_exists"),      True),
            ("Co-Author Name",          data.get("name_of_co_author"),          False),
            ("Co-Author Surname Exists",data.get("co_author_surname_exists"),   True),
            ("Co-Author Surname",       data.get("surname_of_co_author"),       False),
            ("Date Exists",             data.get("date_exists"),                True),
            ("Date",                    data.get("date"),                       False),
        ]
        for i, (label, value, is_bool) in enumerate(fields):
            _data_row(ws, r, label, value, is_bool=is_bool, alt=(i % 2 == 0))
            r += 1
        r += 1


def _build_3d_sheet(ws, results: dict):
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 20

    _header_row(ws, 1, ["Field", "Value", "Status"])
    r = 2

    for form_name, data in results.items():
        _section_row(ws, r, f"📄  {form_name}", 3)
        r += 1

        # Title
        _cell(ws, r, 1, "Title", bold=True, bg=CLR_ROW_ALT)
        _cell(ws, r, 2, data.get("title_name", "—"), bg=CLR_ROW_ALT, wrap=True)
        _bool_cell(ws, r, 3, data.get("title_grammar", False))
        r += 1

        # Abstract
        _cell(ws, r, 1, "Abstract Word Count", bold=True, bg=CLR_ROW_WHITE)
        _cell(ws, r, 2, str(data.get("abstract_word_count", 0)), bg=CLR_ROW_WHITE)
        _bool_cell(ws, r, 3, data.get("abstract_suitable", False))
        r += 1

        # Keywords
        _cell(ws, r, 1, "Keyword Count", bold=True, bg=CLR_ROW_ALT)
        _cell(ws, r, 2, str(data.get("keywords_count", 0)), bg=CLR_ROW_ALT)
        _bool_cell(ws, r, 3, data.get("keywords_suitable", False))
        r += 1

        # Heading numbering
        _cell(ws, r, 1, "Heading Numbering", bold=True, bg=CLR_ROW_WHITE)
        errors = data.get("errors", [])
        err_text = ", ".join(
            [e.get("found", "") if isinstance(e, dict) else str(e) for e in errors]
        ) if errors else "No errors"
        _cell(ws, r, 2, err_text, bg=CLR_ROW_WHITE, wrap=True)
        _bool_cell(ws, r, 3, data.get("numbering_correct", False))
        r += 1

        # References
        ref_data = data.get("references", {})
        _cell(ws, r, 1, "References (Harvard)", bold=True, bg=CLR_ROW_ALT)
        _cell(ws, r, 2, f"Total: {ref_data.get('total_references', 0)}", bg=CLR_ROW_ALT)
        _bool_cell(ws, r, 3, ref_data.get("all_harvard", False))
        r += 1

        # Individual references
        for ref in ref_data.get("references", []):
            _cell(ws, r, 1, "  └ Reference", bg=CLR_ROW_WHITE)
            _cell(ws, r, 2, ref.get("reference", "")[:100], bg=CLR_ROW_WHITE, wrap=True)
            _bool_cell(ws, r, 3, ref.get("is_harvard", False))
            r += 1

        r += 1  # spacer


def _build_summary_sheet(ws, all_results: dict):
    """Overview sheet showing all forms and pass/fail at a glance."""
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    _header_row(ws, 1, ["Form Name", "Type", "Key Check 1", "Key Check 2", "Signature"])
    r = 2

    for form_name, entry in all_results.items():
        form_type = entry.get("form_type", "?")
        data = entry.get("data", {})
        alt = (r % 2 == 0)
        bg = CLR_ROW_ALT if alt else CLR_ROW_WHITE

        _cell(ws, r, 1, form_name, bold=True, bg=bg)
        _cell(ws, r, 2, form_type.upper(), bg=bg, center=True, bold=True)

        if form_type == "3a":
            _bool_cell(ws, r, 3, data.get("corresponding_author_exists", False))
            _bool_cell(ws, r, 4, data.get("title_in_2_exists", False))
            _bool_cell(ws, r, 5, data.get("signature_exists", False))
        elif form_type == "3b":
            _bool_cell(ws, r, 3, data.get("paper_title_exists", False))
            _cell(ws, r, 4, "—", bg=bg, center=True)
            _bool_cell(ws, r, 5, data.get("signature_exists", False))
        elif form_type == "3c":
            _bool_cell(ws, r, 3, data.get("paper_title_exists", False))
            _bool_cell(ws, r, 4, data.get("signatures_exist", False))
            _cell(ws, r, 5, "—", bg=bg, center=True)
        elif form_type == "3d":
            _bool_cell(ws, r, 3, data.get("title_grammar", False))
            _bool_cell(ws, r, 4, data.get("abstract_suitable", False))
            _bool_cell(ws, r, 5, data.get("keywords_suitable", False))
        else:
            for col in range(3, 6):
                _cell(ws, r, col, "—", bg=bg, center=True)
        r += 1


# ── Public API ────────────────────────────────────────────────────────────────

def build_excel(all_results: dict) -> bytes:
    """
    all_results = {
        "filename": {
            "form_type": "3a" | "3b" | "3c" | "3d",
            "data": { ...checker output dict... }
        },
        ...
    }
    Returns bytes of the Excel file.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Group by form type
    grouped = {"3a": {}, "3b": {}, "3c": {}, "3d": {}}
    for fname, entry in all_results.items():
        ft = entry.get("form_type", "3d")
        grouped[ft][fname] = entry.get("data", {})

    # Summary sheet first
    ws_summary = wb.create_sheet("📊 Summary")
    _build_summary_sheet(ws_summary, all_results)

    # Per-type sheets
    builders = {
        "3a": ("📋 Form 3A", _build_3a_sheet),
        "3b": ("📋 Form 3B", _build_3b_sheet),
        "3c": ("📋 Form 3C", _build_3c_sheet),
        "3d": ("📋 Form 3D", _build_3d_sheet),
    }
    for ft, (sheet_name, builder) in builders.items():
        if grouped[ft]:
            ws = wb.create_sheet(sheet_name)
            builder(ws, grouped[ft])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()